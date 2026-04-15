"""
Book Parser Script
==================
Parses book .txt and .json (table of contents) files,
produces output JSON that the HTML viewer can load.

Folder structure:
    books/
        <book-name>/
            *.txt       - book text file
            *.json      - table of contents file
    output/
        <book-name>/
            book-output.json
            book-output.xlsx

Usage:
    python parser.py                       # parse all books in books/
    python parser.py --all                 # same as above
    python parser.py <book-name>           # parse a single book by folder name
    python parser.py <book.txt> <toc.json> # legacy: parse specific files

The .txt file uses these tags:
    <page_number>...</page_number>  - page boundaries and labels
    <heading>...</heading>          - section headings
    <note>...</note>                - footnotes
    <br>                            - line breaks

The .json file can be either:
  1) Plain hierarchy array (legacy):
       [{"chapter": "...", "sections": ["...", ...]}, ...]
  2) Object format with hierarchy + spelling patches:
       {
         "hierarchy": [...],
         "global_patches": {"wrong": "correct", ...},
         "page_patches": {"67": {"wrong": "correct"}, ...}
       }
"""

import json
import os
import re
import sys
import argparse
import glob
from html import unescape

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Ensure Unicode prints work in Windows terminals.
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')


BOOKS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'books')
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')

ARABIC_CHAR_RE = re.compile(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF]')
BENGALI_CHAR_RE = re.compile(r'[\u0980-\u09FF]')
BENGALI_LABEL_HINT_RE = re.compile(r"(দু['’]?আ|দো['’]?আ|দুয়া|দোয়া|আমল|দুরূদ|যিকির|তাসবীহ|সালাম)")
SPECIAL_TAGS = ["meaning", "pronunciation"]
TAG_PATTERN = re.compile(r"^<(" + "|".join(SPECIAL_TAGS) + r")>(.*)</\1>$", re.DOTALL)


def classify_line(line):
    """Classify a line as arabic, empty, or other based on character ratios."""
    stripped = re.sub(r'<[^>]+>', '', line).strip()
    if not stripped:
        return 'empty', line

    arabic_count = 0
    total = len(stripped)
    for ch in stripped:
        if '\u0600' <= ch <= '\u06FF' or '\u0750' <= ch <= '\u077F' or \
           '\uFB50' <= ch <= '\uFDFF' or '\uFE70' <= ch <= '\uFEFF':
            arabic_count += 1

    if arabic_count > 0 and arabic_count / total > 0.6:
        return 'arabic', line.strip()

    return 'other', line


def split_leading_bengali_label(text):
    """
    Extract Bengali label prefixes from mixed lines like:
    "দু'আ ৫: ...Arabic..."
    Returns: (label_prefix, arabic_part)
    """
    stripped = re.sub(r'<[^>]+>', '', text).strip()
    if not stripped:
        return '', stripped

    first_arabic = ARABIC_CHAR_RE.search(stripped)
    if not first_arabic:
        return '', stripped

    prefix = stripped[:first_arabic.start()].strip()
    arabic_part = stripped[first_arabic.start():].strip()
    if not prefix or not arabic_part:
        return '', stripped

    if not BENGALI_CHAR_RE.search(prefix):
        return '', stripped

    looks_like_label = (
        ':' in prefix
        or 'ঃ' in prefix
        or re.search(r'[\u09E6-\u09EF0-9]', prefix) is not None
        or BENGALI_LABEL_HINT_RE.search(prefix) is not None
    )
    if not looks_like_label or len(prefix) > 64:
        return '', stripped

    return prefix, arabic_part


def normalize_toc_data(raw_toc):
    """Normalize chapter hierarchy into [{'chapter': str, 'sections': list[str]}]."""
    if not isinstance(raw_toc, list):
        raise ValueError('Invalid hierarchy format: expected a list')

    normalized = []
    for item in raw_toc:
        if isinstance(item, dict):
            chapter = str(item.get('chapter', '')).strip()
            sections_raw = item.get('sections', [])
        elif isinstance(item, str):
            chapter = item.strip()
            sections_raw = []
        else:
            continue

        if not isinstance(sections_raw, list):
            sections_raw = []
        sections = [str(section).strip() for section in sections_raw]

        normalized.append({
            'chapter': chapter,
            'sections': sections,
        })

    return normalized


def as_replacement_map(value):
    """Normalize replacement map into {old_text: new_text} with string keys/values."""
    if not isinstance(value, dict):
        return {}

    normalized = {}
    for old_text, new_text in value.items():
        if not isinstance(old_text, str):
            old_text = str(old_text)
        if not isinstance(new_text, str):
            new_text = str(new_text)
        if old_text:
            normalized[old_text] = new_text
    return normalized


def parse_label_patches_from_entries(entries):
    """
    Parse spelling entries like:
      [{"label": "67", "m": {"wrong": "correct"}}, ...]
    """
    patches_by_label = {}
    if not isinstance(entries, list):
        return patches_by_label

    for entry in entries:
        if not isinstance(entry, dict):
            continue
        label = str(entry.get('label', '')).strip()
        if not label:
            continue
        replacement_map = as_replacement_map(entry.get('m'))
        if replacement_map:
            patches_by_label[label] = replacement_map
    return patches_by_label


def load_toc_and_spelling_config(json_path):
    """Load TOC hierarchy + spelling patches from either legacy or new JSON format."""
    with open(json_path, 'r', encoding='utf-8') as f:
        raw_data = json.load(f)

    if isinstance(raw_data, list):
        return normalize_toc_data(raw_data), {'global': {}, 'by_label': {}}

    if not isinstance(raw_data, dict):
        raise ValueError(f'Unsupported JSON format in {json_path}')

    raw_toc = raw_data.get('hierarchy')
    if raw_toc is None:
        raw_toc = raw_data.get('toc')
    if raw_toc is None:
        raise ValueError(
            f"JSON file '{json_path}' must be a hierarchy list, or object with 'hierarchy'/'toc'"
        )

    global_patches = as_replacement_map(raw_data.get('global_patches'))
    by_label = {}

    for key in ('page_patches', 'patches_by_label'):
        value = raw_data.get(key)
        if isinstance(value, dict):
            for label, patch_map in value.items():
                normalized_map = as_replacement_map(patch_map)
                if normalized_map:
                    by_label[str(label)] = normalized_map

    # Optional compatibility with spelling list structure:
    # {"spelling_fixes": [{"label": "...", "m": {...}}, ...]}
    by_label.update(parse_label_patches_from_entries(raw_data.get('spelling_fixes')))

    # Optional compatibility with compact label->map structure:
    # {"spelling_fixes": {"67": {"wrong": "correct"}}}
    if isinstance(raw_data.get('spelling_fixes'), dict):
        for label, patch_map in raw_data['spelling_fixes'].items():
            normalized_map = as_replacement_map(patch_map)
            if normalized_map:
                by_label[str(label)] = normalized_map

    return normalize_toc_data(raw_toc), {'global': global_patches, 'by_label': by_label}


def strip_html_tags(value):
    """Remove tags from text for plain-content replacements."""
    return re.sub(r'<[^>]+>', '', value)


def apply_spelling_patches(pages, patch_config):
    """
    Apply spelling patches after section hierarchy is built.
    Supports:
      - global patches for all pages
      - page-label specific patches
    """
    global_patches = as_replacement_map(patch_config.get('global'))
    by_label = patch_config.get('by_label', {})

    if not global_patches and not by_label:
        return {'replacements': 0, 'tags_added': 0}

    total_replacements = 0
    total_tags_added = 0
    global_hits = {old_text: 0 for old_text in global_patches}

    for page in pages:
        label = str(page.get('label', ''))
        label_candidates = {label}
        for original_label in page.get('originalLabels', []):
            label_candidates.add(str(original_label))

        label_patches = {}
        for candidate in label_candidates:
            if candidate in by_label:
                label_patches.update(as_replacement_map(by_label[candidate]))

        html = page.get('html', '')
        plain_content = page.get('plainContent', '')
        plain_notes = page.get('plainNotes', '')
        section_heading = page.get('sectionHeading', '')

        page_replacements = 0
        page_tags = 0

        for old_text, new_text in global_patches.items():
            match_count = html.count(old_text)
            if match_count <= 0:
                continue

            html = html.replace(old_text, new_text)
            global_hits[old_text] += match_count
            page_replacements += match_count

            if TAG_PATTERN.match(new_text.strip()):
                page_tags += match_count

            old_plain = strip_html_tags(old_text)
            new_plain = strip_html_tags(new_text)
            if old_plain:
                plain_content = plain_content.replace(old_plain, new_plain)
                plain_notes = plain_notes.replace(old_plain, new_plain)
                section_heading = section_heading.replace(old_plain, new_plain)

        for old_text, new_text in label_patches.items():
            match_count = html.count(old_text)
            if match_count <= 0:
                preview = old_text[:60] + ('...' if len(old_text) > 60 else '')
                print(f"  [WARNING] Label patch not found in page '{label}': {preview}")
                continue

            html = html.replace(old_text, new_text)
            page_replacements += match_count

            if TAG_PATTERN.match(new_text.strip()):
                page_tags += match_count

            old_plain = strip_html_tags(old_text)
            new_plain = strip_html_tags(new_text)
            if old_plain:
                plain_content = plain_content.replace(old_plain, new_plain)
                plain_notes = plain_notes.replace(old_plain, new_plain)
                section_heading = section_heading.replace(old_plain, new_plain)

        if page_replacements > 0:
            tag_info = f' ({page_tags} tags added)' if page_tags else ''
            print(f"  Spelling page '{label}': {page_replacements} replacements{tag_info}")

        page['html'] = html
        page['plainContent'] = plain_content
        page['plainNotes'] = plain_notes
        page['sectionHeading'] = section_heading

        total_replacements += page_replacements
        total_tags_added += page_tags

    missing_globals = [old_text for old_text, hits in global_hits.items() if hits == 0]
    if missing_globals:
        print(f'  [WARNING] {len(missing_globals)} global patch(es) were not found in parsed pages')

    return {
        'replacements': total_replacements,
        'tags_added': total_tags_added,
    }


def format_content(raw):
    """Convert raw page content into styled HTML."""
    html = raw

    # Convert <heading> tags
    def replace_heading(m):
        return f'<div class="heading">{m.group(1).strip()}</div>'

    html = re.sub(r'<heading>([\s\S]*?)</heading>', replace_heading, html, flags=re.IGNORECASE)
    # Keep headings separate so they never get merged into an arabic-text block.
    html = re.sub(
        r'\s*(<div class="heading">[\s\S]*?</div>)\s*',
        r'<br>\1<br>',
        html,
        flags=re.IGNORECASE,
    )

    # Split by <br> and classify lines
    lines = html.split('<br>')
    classified = [classify_line(line) for line in lines]

    # Merge consecutive arabic lines into one block
    processed = []
    arabic_buffer = []

    for idx, (line_type, content) in enumerate(classified):
        if line_type == 'arabic':
            label_prefix, arabic_only = split_leading_bengali_label(content)
            if label_prefix:
                if arabic_buffer:
                    processed.append(f'<div class="arabic-text">{" ".join(arabic_buffer)}</div>')
                    arabic_buffer = []
                processed.append(f'<div class="arabic-inline-label">{label_prefix}</div>')
                content = arabic_only
            arabic_buffer.append(content)
            continue
        # If Arabic lines are separated only by blank lines, keep them together
        # in one arabic-text block instead of splitting into multiple boxes.
        if line_type == 'empty' and arabic_buffer:
            next_type = None
            for j in range(idx + 1, len(classified)):
                if classified[j][0] != 'empty':
                    next_type = classified[j][0]
                    break
            if next_type == 'arabic':
                continue
        if arabic_buffer:
            processed.append(f'<div class="arabic-text">{" ".join(arabic_buffer)}</div>')
            arabic_buffer = []
        processed.append(content)

    if arabic_buffer:
        processed.append(f'<div class="arabic-text">{" ".join(arabic_buffer)}</div>')

    html = '<br>'.join(processed)

    # Clean up excessive line breaks
    html = re.sub(r'(<br>\s*){3,}', '<br><br>', html)
    html = re.sub(r'^(<br>\s*)+', '', html)

    return f'<div class="page-text">{html}</div>'


def split_page_content_by_heading(content):
    """Split a page into chunks where each <heading> starts a new section chunk."""
    heading_pattern = re.compile(r'<heading>[\s\S]*?</heading>', re.IGNORECASE)
    matches = list(heading_pattern.finditer(content))

    if not matches:
        return [{'hasHeading': False, 'content': content}] if content.strip() else []

    chunks = []
    cursor = 0
    for i, match in enumerate(matches):
        leading = content[cursor:match.start()]
        if leading.strip():
            chunks.append({'hasHeading': False, 'content': leading})

        next_start = matches[i + 1].start() if i + 1 < len(matches) else len(content)
        section_chunk = content[match.start():next_start]
        if section_chunk.strip():
            chunks.append({'hasHeading': True, 'content': section_chunk})

        cursor = next_start

    trailing = content[cursor:]
    if trailing.strip():
        chunks.append({'hasHeading': False, 'content': trailing})

    return chunks


def extract_notes(content):
    """Extract <note> blocks from content and return cleaned text and note list."""
    notes = []

    def collect_note(m):
        notes.append(m.group(1).strip())
        return ''

    cleaned = re.sub(r'<note>([\s\S]*?)</note>', collect_note, content, flags=re.IGNORECASE)
    return cleaned, notes


def parse_book(text):
    """Parse the raw text file into section-based pages."""
    # Split by <page_number> tags
    raw_parts = re.split(r'<page_number>', text, flags=re.IGNORECASE)

    raw_pages = []
    for i in range(1, len(raw_parts)):
        chunk = raw_parts[i]
        close_idx = chunk.find('</page_number>')
        if close_idx == -1:
            continue

        page_num_raw = chunk[:close_idx].strip()
        content = chunk[close_idx + len('</page_number>'):].strip()

        # Extract page label (Bengali or ASCII digits)
        num_match = re.search(r'[\u09E6-\u09EF\d]+', page_num_raw)
        page_label = num_match.group(0) if num_match else page_num_raw

        raw_pages.append({
            'label': page_label,
            'fullLabel': page_num_raw,
            'content': content,
        })

    # Group chunks into sections.
    # A section starts at each heading, even when multiple headings exist
    # inside the same original page.
    sections = []
    current_section = None

    for page in raw_pages:
        chunks = split_page_content_by_heading(page['content'])
        for chunk in chunks:
            chunk_content, chunk_notes = extract_notes(chunk['content'])

            if chunk['hasHeading'] or current_section is None:
                current_section = {'parts': [], 'notes': [], 'labels': []}
                sections.append(current_section)

            if chunk_content.strip():
                current_section['parts'].append(chunk_content)
            current_section['notes'].extend(chunk_notes)

            if not current_section['labels'] or current_section['labels'][-1] != page['label']:
                current_section['labels'].append(page['label'])

    # Build one combined page per section
    result = []
    for section in sections:
        combined_html = ''
        raw_texts = section['parts']
        all_notes = section['notes']

        # Format the entire section in one pass so long Arabic quotations are not
        # split into separate blocks when source page boundaries occur mid-quote.
        if raw_texts:
            combined_html = format_content(''.join(raw_texts))

        # Append footnotes
        if all_notes:
            combined_html += (
                '<div class="section-footnotes">'
                '<div class="section-footnotes-label">\u09a4\u09a5\u09cd\u09af\u09b8\u09c2\u09a4\u09cd\u09b0</div>'
                f'<div class="note">{"<br>".join(all_notes)}</div>'
                '</div>'
            )

        # Page label: range if section spans multiple original pages
        original_labels = section['labels']
        first_label = original_labels[0]
        last_label = original_labels[-1]
        label = first_label if first_label == last_label else f'{first_label}-{last_label}'

        # Extract plain text content (strip heading and HTML tags)
        raw_content = '\n'.join(raw_texts)
        raw_content_no_heading = re.sub(r'<heading>[\s\S]*?</heading>', '', raw_content, flags=re.IGNORECASE)
        plain_content = re.sub(r'<[^>]+>', '', raw_content_no_heading).strip()
        plain_content = re.sub(r'\n{3,}', '\n\n', plain_content)

        # Extract heading text
        heading_match = re.search(r'<heading>([\s\S]*?)</heading>', raw_content, re.IGNORECASE)
        section_heading = heading_match.group(1).strip() if heading_match else ''

        plain_notes = '\n'.join(all_notes)
        plain_notes = re.sub(r'<[^>]+>', '', plain_notes).strip()

        result.append({
            'label': label,
            'originalLabels': original_labels,
            'html': combined_html,
            'plainContent': plain_content,
            'plainNotes': plain_notes,
            'sectionHeading': section_heading,
        })

    return result

def detect_files_in_folder(folder_path):
    """Detect .txt and .json files inside a book folder."""
    txt_files = glob.glob(os.path.join(folder_path, '*.txt'))
    json_files = [f for f in glob.glob(os.path.join(folder_path, '*.json'))
                  if not os.path.basename(f).startswith('book-output')]

    if not txt_files:
        raise FileNotFoundError(f'No .txt file found in {folder_path}')
    if not json_files:
        raise FileNotFoundError(f'No .json file found in {folder_path}')

    if len(txt_files) > 1:
        print(f'  Warning: Multiple .txt files in {folder_path}, using first: {os.path.basename(txt_files[0])}')
    if len(json_files) > 1:
        print(f'  Warning: Multiple .json files in {folder_path}, using first: {os.path.basename(json_files[0])}')

    return txt_files[0], json_files[0]


def discover_books():
    """Find all book folders inside the books/ directory."""
    if not os.path.isdir(BOOKS_DIR):
        print(f'Error: books/ directory not found at {BOOKS_DIR}')
        print('Create books/<book-name>/ folders with .txt and .json files.')
        sys.exit(1)

    book_folders = []
    for entry in sorted(os.listdir(BOOKS_DIR)):
        full_path = os.path.join(BOOKS_DIR, entry)
        if os.path.isdir(full_path):
            book_folders.append((entry, full_path))

    if not book_folders:
        print('Error: No book folders found in books/')
        print('Create books/<book-name>/ folders with .txt and .json files.')
        sys.exit(1)

    return book_folders


def generate_xlsx(toc_data, pages, output_path):
    """Generate an xlsx file that mirrors the viewer's XLSX download output."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Book Content'

    # Header row
    headers = ['chapter_id', 'chapter_name', 'section_id', 'section_name', 'content', 'notes']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    def normalize_title(value):
        return re.sub(r'\s+', ' ', str(value or '')).strip()

    def find_div_block_bounds(text, start_idx):
        """Return (start, end) for a <div>...</div> block starting at start_idx."""
        token_re = re.compile(r'<div\b[^>]*>|</div>', flags=re.IGNORECASE)
        depth = 0
        block_start = None

        for match in token_re.finditer(text, pos=start_idx):
            token = match.group(0).lower()
            if token.startswith('<div'):
                if depth == 0:
                    block_start = match.start()
                depth += 1
            else:
                depth -= 1
                if depth == 0 and block_start is not None:
                    return block_start, match.end()
                if depth < 0:
                    break

        return None

    def extract_tagged_from_html(html_text):
        """Match the browser XLSX extraction: keep content HTML, notes as plain text."""
        source_html = html_text or ''
        notes_text = ''

        footnotes_start = re.search(
            r'<div\s+class=["\']section-footnotes["\'][^>]*>',
            source_html,
            flags=re.IGNORECASE,
        )
        if footnotes_start:
            bounds = find_div_block_bounds(source_html, footnotes_start.start())
            if bounds:
                start, end = bounds
                footnotes_html = source_html[start:end]
                note_match = re.search(
                    r'<div\s+class=["\']note["\'][^>]*>([\s\S]*?)</div>',
                    footnotes_html,
                    flags=re.IGNORECASE,
                )
                if note_match:
                    note_html = note_match.group(1)
                    note_html = re.sub(r'<br\s*/?>', '\n', note_html, flags=re.IGNORECASE)
                    notes_text = unescape(re.sub(r'<[^>]+>', '', note_html)).strip()
                source_html = source_html[:start] + source_html[end:]

        source_html = re.sub(
            r'<div\s+class=["\']heading["\'][^>]*>[\s\S]*?</div>',
            '',
            source_html,
            flags=re.IGNORECASE,
        )

        return {
            'content': source_html.strip(),
            'notes': notes_text,
        }

    safe_toc = [ch for ch in toc_data if isinstance(ch, dict)] if isinstance(toc_data, list) else []
    safe_pages = [pg for pg in pages if isinstance(pg, dict)] if isinstance(pages, list) else []

    def add_styled_export_row(row_data):
        ws.append([
            row_data.get('chapter_id', ''),
            row_data.get('chapter_name', ''),
            row_data.get('section_id', ''),
            row_data.get('section_name', ''),
            row_data.get('content', ''),
            row_data.get('notes', ''),
        ])
        row_number = ws.max_row
        for col_number in range(1, 7):
            cell = ws.cell(row=row_number, column=col_number)
            cell.border = thin_border
            if col_number in (5, 6):
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    chapter_meta = []
    for ch_idx, chapter in enumerate(safe_toc):
        chapter_name = str(chapter.get('chapter', '')).strip()
        raw_sections = chapter.get('sections', [])
        if isinstance(raw_sections, list):
            normalized_sections = [str(sec or '').strip() for sec in raw_sections]
            normalized_sections = [sec for sec in normalized_sections if sec]
        else:
            normalized_sections = []

        sections = normalized_sections or [chapter_name or f'Chapter {ch_idx + 1}']
        chapter_meta.append({
            'chapter_name': chapter_name,
            'raw_sections': normalized_sections,
            'sections': sections,
        })

    # Map section/chapter headings to chapter index so page grouping follows UI behavior.
    heading_to_chapter_index = {}
    for ch_idx, meta in enumerate(chapter_meta):
        chapter_norm = normalize_title(meta['chapter_name'])
        if chapter_norm and chapter_norm not in heading_to_chapter_index:
            heading_to_chapter_index[chapter_norm] = ch_idx
        for section_name in meta['raw_sections']:
            section_norm = normalize_title(section_name)
            if section_norm and section_norm not in heading_to_chapter_index:
                heading_to_chapter_index[section_norm] = ch_idx

    chapter_pages = [[] for _ in chapter_meta]
    current_chapter_idx = 0 if chapter_meta else -1
    for page in safe_pages:
        heading_norm = normalize_title(page.get('sectionHeading', ''))
        if heading_norm and heading_norm in heading_to_chapter_index:
            current_chapter_idx = heading_to_chapter_index[heading_norm]

        if 0 <= current_chapter_idx < len(chapter_pages):
            chapter_pages[current_chapter_idx].append(page)

    row_count = 0
    for ch_idx, meta in enumerate(chapter_meta):
        pages_for_chapter = chapter_pages[ch_idx] if ch_idx < len(chapter_pages) else []
        page_cursor = 0

        for sec_idx, section_name in enumerate(meta['sections']):
            section_pages = []
            if len(meta['raw_sections']) == 0:
                # No TOC sections: keep one synthesized section with all chapter content.
                section_pages = pages_for_chapter[page_cursor:]
                page_cursor = len(pages_for_chapter)
            else:
                # One row per section; any overflow pages are appended to last section.
                if page_cursor < len(pages_for_chapter):
                    section_pages.append(pages_for_chapter[page_cursor])
                    page_cursor += 1
                if sec_idx == len(meta['sections']) - 1 and page_cursor < len(pages_for_chapter):
                    section_pages.extend(pages_for_chapter[page_cursor:])
                    page_cursor = len(pages_for_chapter)

            content = ''
            notes = ''
            if section_pages:
                extracted_parts = [extract_tagged_from_html(page.get('html', '')) for page in section_pages]
                content = '\n\n'.join(part['content'] for part in extracted_parts if part['content'])
                notes = '\n\n'.join(part['notes'] for part in extracted_parts if part['notes'])

            add_styled_export_row({
                'chapter_id': ch_idx + 1,
                'chapter_name': meta['chapter_name'],
                'section_id': sec_idx + 1,
                'section_name': section_name,
                'content': content,
                'notes': notes,
            })
            row_count += 1

    # Fallback for direct/manual JSON with no usable TOC rows.
    if row_count == 0:
        for i, page in enumerate(safe_pages):
            extracted = extract_tagged_from_html(page.get('html', ''))
            section_heading = str(page.get('sectionHeading', '') or '').strip()
            label = str(page.get('label', '') or '').strip()
            fallback_name = section_heading or f'Page {label or i + 1}'

            add_styled_export_row({
                'chapter_id': '',
                'chapter_name': '',
                'section_id': i + 1,
                'section_name': fallback_name,
                'content': extracted['content'],
                'notes': extracted['notes'],
            })
            row_count += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 40

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return row_count


def generate_books_index(output_dir):
    """Generate a books-index.json listing all available parsed books."""
    books = []
    for entry in sorted(os.listdir(output_dir)):
        book_output_dir = os.path.join(output_dir, entry)
        json_path = os.path.join(book_output_dir, 'book-output.json')
        if os.path.isdir(book_output_dir) and os.path.isfile(json_path):
            books.append({
                'id': entry,
                'name': entry.replace('-', ' ').title(),
                'path': f'output/{entry}/book-output.json',
            })

    index_path = os.path.join(output_dir, 'books-index.json')
    with open(index_path, 'w', encoding='utf-8') as f:
        json.dump(books, f, ensure_ascii=False, indent=2)

    return books


def parse_single_book(book_name, book_folder):
    """Parse a single book and write output files."""
    print(f'\n--- Parsing: {book_name} ---')

    txt_file, json_file = detect_files_in_folder(book_folder)
    print(f'  Text: {os.path.basename(txt_file)}')
    print(f'  TOC:  {os.path.basename(json_file)}')

    # Read input files
    with open(txt_file, 'r', encoding='utf-8') as f:
        text_data = f.read()

    toc_data, patch_config = load_toc_and_spelling_config(json_file)

    # Parse the book
    pages = parse_book(text_data)
    patch_stats = apply_spelling_patches(pages, patch_config)

    # Build JSON output
    json_pages = []
    for pg in pages:
        json_pages.append({
            'label': pg['label'],
            'originalLabels': pg['originalLabels'],
            'html': pg['html'],
            'sectionHeading': pg['sectionHeading'],
        })

    output = {
        'toc': toc_data,
        'pages': json_pages,
    }

    # Create output directory
    book_output_dir = os.path.join(OUTPUT_DIR, book_name)
    os.makedirs(book_output_dir, exist_ok=True)

    # Write JSON output
    json_output_path = os.path.join(book_output_dir, 'book-output.json')
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f'  Parsed {len(pages)} pages')
    print(f'  TOC: {len(toc_data)} chapters')
    if patch_stats['replacements'] > 0:
        print(f"  Spelling fixes: {patch_stats['replacements']} replacements ({patch_stats['tags_added']} tags)")
    print(f'  JSON -> output/{book_name}/book-output.json')

    # Generate xlsx file
    xlsx_path = os.path.join(book_output_dir, 'book-output.xlsx')
    try:
        xlsx_rows = generate_xlsx(toc_data, pages, xlsx_path)
        print(f'  XLSX -> output/{book_name}/book-output.xlsx ({xlsx_rows} rows)')
    except PermissionError:
        print(f'  Warning: Could not write output/{book_name}/book-output.xlsx (file is open).')
        print('  Close the file and run parser again to refresh XLSX output.')

    return True


def main():
    parser = argparse.ArgumentParser(
        description='Parse book .txt and .json files into output JSON for the HTML viewer.',
        epilog='''
Examples:
  python parser.py                    Parse all books in books/
  python parser.py --all              Same as above
  python parser.py hadis-e-bornito    Parse a single book by folder name
  python parser.py book.txt toc.json  Legacy mode: parse specific files
        ''',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('args', nargs='*', help='Book name, or legacy .txt and .json file paths')
    parser.add_argument('--all', action='store_true', help='Parse all books in books/ directory')
    parser.add_argument('-o', '--output', default=None,
                        help='Output JSON file path (legacy mode only)')

    parsed = parser.parse_args()

    # Legacy mode: two file arguments provided (both existing files)
    if len(parsed.args) == 2 and os.path.isfile(parsed.args[0]) and os.path.isfile(parsed.args[1]):
        txt_file, json_file = parsed.args
        output_path = parsed.output or 'book-output.json'

        with open(txt_file, 'r', encoding='utf-8') as f:
            text_data = f.read()
        toc_data, patch_config = load_toc_and_spelling_config(json_file)

        pages = parse_book(text_data)
        patch_stats = apply_spelling_patches(pages, patch_config)

        json_pages = [
            {
                'label': pg['label'],
                'originalLabels': pg['originalLabels'],
                'html': pg['html'],
                'sectionHeading': pg['sectionHeading'],
            }
            for pg in pages
        ]
        output = {'toc': toc_data, 'pages': json_pages}

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        print(f'Parsed {len(pages)} pages -> {output_path}')
        if patch_stats['replacements'] > 0:
            print(f"Spelling fixes: {patch_stats['replacements']} replacements ({patch_stats['tags_added']} tags)")

        xlsx_path = os.path.splitext(output_path)[0] + '.xlsx'
        try:
            xlsx_rows = generate_xlsx(toc_data, pages, xlsx_path)
            print(f'XLSX -> {xlsx_path} ({xlsx_rows} rows)')
        except PermissionError:
            print(f'Warning: Could not write {xlsx_path} (file is open).')
            print('Close the file and run parser again to refresh XLSX output.')
        return

    # Single book by name
    if len(parsed.args) == 1 and not parsed.all:
        book_name = parsed.args[0]
        book_folder = os.path.join(BOOKS_DIR, book_name)
        if not os.path.isdir(book_folder):
            print(f'Error: Book folder not found: books/{book_name}/')
            print(f'Available books:')
            for name, _ in discover_books():
                print(f'  - {name}')
            sys.exit(1)

        parse_single_book(book_name, book_folder)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        books = generate_books_index(OUTPUT_DIR)
        print(f'\nBooks index updated: {len(books)} book(s)')
        return

    # Default: parse all books
    book_folders = discover_books()
    print(f'Found {len(book_folders)} book(s) in books/')

    success_count = 0
    for book_name, book_folder in book_folders:
        try:
            parse_single_book(book_name, book_folder)
            success_count += 1
        except FileNotFoundError as e:
            print(f'  Skipping {book_name}: {e}')
        except Exception as e:
            print(f'  Error parsing {book_name}: {e}')

    # Generate books index
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    books = generate_books_index(OUTPUT_DIR)
    print(f'\n=== Done: {success_count}/{len(book_folders)} books parsed ===')
    print(f'Books index: output/books-index.json ({len(books)} book(s))')


if __name__ == '__main__':
    main()

